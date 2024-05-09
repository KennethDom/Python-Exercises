# -*- coding: utf-8 -*-
"""
Created on Fri May 14 09:04:07 2021

@author: Usuario
"""

import kivy

import openpyxl
from kivy.app import App
from kivy.uix import button, scrollview, gridlayout, boxlayout
from kivy.app import runTouchApp

class Aplication(App):
    gestor = boxlayout.BoxLayout(orientation = 'horizontal')
    cuadricula = gridlayout.GridLayout(cols=1, size_hint=(None, None),
                width=400, height=5500)

    def extraerPreguntas(self, materia):
        print(materia)
        wb= openpyxl.load_workbook('archivo.xlsx') #cambia al nombre del archivo
        sheet=wb[materia]
        datos = []
        for renglon in range(1, sheet.max_row + 1):
            pregunta = sheet.cell(row=renglon, column=1).value
            datos.append(pregunta)
        return self.agregarPreguntas(datos)

    def agregarPreguntas(self, datos):
        self.cuadricula.clear_widgets()
        for pregunta in datos:
            pregunta = button.Button(text=pregunta, size=(400, 40),
                        size_hint=(1, None))
            self.cuadricula.add_widget(pregunta)

    def build(self):
        gestorMaterias = gridlayout.GridLayout(cols=1, size_hint=(None, None),
                    width=400, height=1000)
        materias = ['Cross-plataform', 'Kivy', 'Python en App Movil', 'Ventajas Kivy', 'Arquitectura Kivy', 'Core']
        materias_d={}
        for mat in materias:
            materias_d[mat]= mat
        print(materias_d)

        for k, v in materias_d.items():
            k = button.Button(text=v, size=(480,40), size_hint=(1, None))
            k.bind(on_press=lambda k:self.extraerPreguntas(k.text))
            gestorMaterias.add_widget(k)
        root=boxlayout.BoxLayout()
        scrollMateria=scrollview.ScrollView(size_hint=(1, 1))
        scrollPregunta=scrollview.ScrollView(size_hint=(1, 1))
        root.add_widget(self.gestor)
        self.gestor.add_widget(scrollMateria)
        scrollMateria.add_widget(gestorMaterias)
        self.gestor.add_widget(scrollPregunta)
        scrollPregunta.add_widget(self.cuadricula)
        return root

Aplication().run